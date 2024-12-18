# -*- coding: utf-8 -*-
from openpyxl import load_workbook


def _determine_start_row_by_given_string(worksheet, string_for_start_row):
    """
    Determines the start row by a given string
    :param worksheet: the worksheet to be searched
    :param string_for_start_row: the string to be searched for as unique content of a cell
    :return: the start row
    """
    start_row = 0
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        for cell in row:
            cell_value = cell.value
            if cell_value == string_for_start_row:
                start_row = cell.row
                break
        if start_row:
            break
    return start_row


def read_from_excel_file(file_path, string_for_start_row):
    """
    Reads from an Excel file and returns dictionaries by section and by file
    :param file_path:
    :param string_for_start_row:
    :return:

    {'A DRAWINGS': {
        'MC077-022-001 LEAK PROOF JOINT DESIGN AND DRAWING FOR TANK AND DECK SURFACE': {
            'line number': 'A.1',
            'file number': 'MC077-022-001',
            'file name': 'LEAK PROOF JOINT DESIGN AND DRAWING FOR TANK AND DECK SURFACE'},
        'MC077-022-002 PIPE SUPPORT, WALKWAY STRUCTURE EXECUTION DRAWING': {
            'line number': 'A.2',
            'file number': 'MC077-022-002',
            'file name': 'PIPE SUPPORT, WALKWAY STRUCTURE EXECUTION DRAWING'}},}

    {'MC077-022-001 LEAK PROOF JOINT DESIGN AND DRAWING FOR TANK AND DECK SURFACE': {
        'section number': 'A DRAWINGS',
        'line number': 'A.1'},
    'MC077-022-002 PIPE SUPPORT, WALKWAY STRUCTURE EXECUTION DRAWING': {
        'section number': 'A DRAWINGS',
        'line number': 'A.2'},}
    """
    workbook = load_workbook(file_path)
    worksheet = workbook[workbook.sheetnames[0]]
    dict_by_section = {}
    dict_by_file = {}

    # determine start row
    start_row = _determine_start_row_by_given_string(worksheet, string_for_start_row)
    if not start_row:
        return 'Error: Could not find the start row', None

    # reading
    for row in worksheet.iter_rows(min_row=start_row, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        line_number = row[0].value      # A.1 or A
        drawing_number = row[1].value   # M077-021-001 or None
        drawing_name = row[2].value     # Hull... or DRAWINGS

        # ignore without line_number
        if not line_number:
            continue

        # check if it is a section
        if line_number and not drawing_number and drawing_name:

            # stripping
            line_number = line_number.strip()
            drawing_name = drawing_name.strip()

            section_number_space_name = line_number + ' ' + drawing_name
            dict_by_section[section_number_space_name] = {}
            continue

        # check if it is a file
        if line_number and drawing_number and drawing_name:

            # stripping
            line_number = line_number.strip()
            drawing_number = drawing_number.strip()
            drawing_name = drawing_name.strip()

            section_number = line_number[0]

            section_number_space_name = None
            for key in dict_by_section.keys():
                if section_number == key[0]:
                    section_number_space_name = key
                    break

            dict_by_section[section_number_space_name][drawing_number + ' ' + drawing_name] = {}
            dict_by_section[section_number_space_name][drawing_number + ' ' + drawing_name]['line number'] = line_number
            dict_by_section[section_number_space_name][drawing_number + ' ' + drawing_name]['file number'] = drawing_number
            dict_by_section[section_number_space_name][drawing_number + ' ' + drawing_name]['file name'] = drawing_name

    # generating dict3
    for section_number_space_name in dict_by_section.keys():
        for file_number_space_name in dict_by_section[section_number_space_name].keys():
            dict_by_file[file_number_space_name] = {}
            dict_by_file[file_number_space_name]['section number'] = section_number_space_name
            dict_by_file[file_number_space_name]['line number'] = dict_by_section[section_number_space_name][file_number_space_name]['line number']

    # returning the dictionary
    return dict_by_section, dict_by_file
